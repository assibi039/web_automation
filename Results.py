import openpyxl
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys

wb = load_workbook('tg.xlsx')
wb2 = load_workbook('test.xlsx')

sheet = wb['CS']
sheet2 = wb2['Sheet1']

driver = webdriver.Chrome(executable_path='D:\chromedriver_win32\chromedriver.exe')
driver.get('https://web.karunya.edu/exam/login_may20.asp')

j=2

for i in range(3,349):
    try:
        print(i)
        elem = driver.find_element_by_xpath('/html/body/table/tbody/tr/td/form/div/table/tbody/tr/td/p[1]/input')
        elem.send_keys(sheet.cell(row=i,column=2).value)

        elem = driver.find_element_by_xpath('/html/body/table/tbody/tr/td/form/div/table/tbody/tr/td/p[3]/input')
        elem.send_keys(sheet.cell(row=i,column=5).value.date().strftime("%d%m%Y"))

        elem = driver.find_element_by_xpath('/html/body/table/tbody/tr/td/form/div/table/tbody/tr/td/p[8]/input[1]')
        elem.click()

        elem = driver.find_element_by_xpath('/html/body/table[2]/tbody/tr[2]/td[1]').text
        sheet2.cell(row=j,column=1).value=elem

        elem = driver.find_element_by_xpath('/html/body/table[2]/tbody/tr[2]/td[2]').text
        sheet2.cell(row=j,column=2).value=elem
        print(elem)

        elem = driver.find_element_by_xpath('/html/body/center[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td[2]/font/b').text
        sheet2.cell(row=j,column=3).value=elem

        elem = driver.find_element_by_xpath('/html/body/center[2]/table/tbody/tr/td[3]/table/tbody/tr[2]/td/font/b').text
        sheet2.cell(row=j,column=4).value=elem

        elem = driver.find_element_by_xpath('/html/body/table[1]/tbody/tr/td[2]/p/a/b')
        elem.click()
        j=j+1
        wb2.save('test.xlsx')

        
    except Exception as ex:
        print("invalid")
        elem = driver.find_element_by_xpath('/html/body/table[1]/tbody/tr/td[2]/p/a/b')
        elem.click()
        wb2.save('test.xlsx')

        
wb2.save('test.xlsx')
driver.quit
